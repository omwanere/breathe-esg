import csv
import io
import math
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from django.db import transaction
from django.utils import timezone
from apps.ingestion.models import IngestionJob, DataSource, UnitConversion, EmissionFactor
from apps.review.models import RawActivityRow
from apps.export.models import AuditLog

# Airport Coordinates Lookup for Great-Circle Distance
AIRPORT_COORDINATES = {
    'ATL': (33.6407, -84.4277),
    'PEK': (40.0799, 116.6031),
    'LAX': (33.9416, -118.4085),
    'HND': (35.5494, 139.7798),
    'DXB': (25.2532, 55.3657),
    'ORD': (41.9742, -87.9073),
    'LHR': (51.4700, -0.4543),
    'PVG': (31.1443, 121.8083),
    'CDG': (49.0097, 2.5479),
    'DFW': (32.8998, -97.0403),
    'AMS': (52.3105, 4.7683),
    'HKG': (22.3080, 113.9185),
    'FRA': (50.0379, 8.5622),
    'DEN': (39.8561, -104.6737),
    'DEL': (28.5562, 77.1000),
    'SIN': (1.3644, 103.9915),
    'BOM': (19.0896, 72.8656),
    'SYD': (-33.9461, 151.1772),
    'LAX': (33.9416, -118.4085),
    'SFO': (37.6190, -122.3749),
    'JFK': (40.6398, -73.7789),
    'SEA': (47.4502, -122.3088),
    'MIA': (25.7959, -80.2871),
    'YYZ': (43.6777, -79.6248),
    'YVR': (49.1967, -123.1815),
    'MEX': (19.4361, -99.0719),
    'GRU': (-23.4356, -46.4731),
    'EZE': (-34.8222, -58.5358),
    'CPT': (-33.9711, 18.6017),
    'JNB': (-26.1392, 28.2460),
    'CAI': (30.1219, 31.4056),
    'IST': (41.2753, 28.7519),
    'ATH': (37.9356, 23.9445),
    'BKK': (13.6900, 100.7501),
    'ICN': (37.4602, 126.4407),
    'KUL': (2.7456, 101.7099),
    'CGK': (-6.1256, 106.6559),
    'MNL': (14.5086, 121.0194),
    'TPE': (25.0797, 121.2342),
    'CAN': (23.3924, 113.2988),
    'MEL': (-37.6690, 144.8410),
    'AKL': (-37.0081, 174.7917),
    'DOH': (25.2611, 51.5650),
    'RUH': (24.9576, 46.6988),
    'ZRH': (47.4582, 8.5555),
    'GVA': (46.2381, 6.1089),
    'ARN': (59.6519, 17.9186),
    'OSL': (60.1976, 11.1004),
    'CPH': (55.6180, 12.6560),
    'DUB': (53.4213, -6.2701),
    'LIS': (38.7742, -9.1342),
}

def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371.0  # Earth radius in kilometers
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    
    a = math.sin(delta_phi / 2.0) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2.0) ** 2
    c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))
    return R * c

def read_file_data(file_obj, filename):
    """
    Reads CSV or Excel and yields dictionary rows.
    Handles encoding issues gracefully (tries utf-8, fallbacks to latin-1).
    """
    if filename.endswith('.xlsx'):
        wb = load_workbook(file_obj, data_only=True)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # skip empty rows
                continue
            yield dict(zip(headers, row))
    else:
        # It's a CSV. Try UTF-8 first, fallback to Latin-1
        try:
            content = file_obj.read().decode('utf-8')
        except UnicodeDecodeError:
            file_obj.seek(0)
            content = file_obj.read().decode('latin-1')
        
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            yield row

def parse_german_decimal(val):
    """Converts a German decimal format (e.g. 1.234,56 or 1234,56) to Decimal."""
    if val is None:
        return Decimal('0')
    val_str = str(val).strip()
    if not val_str:
        return Decimal('0')
    if ',' in val_str:
        val_str = val_str.replace('.', '').replace(',', '.')
    return Decimal(val_str)

def parse_standard_date(val, formats=('%Y-%m-%d', '%d.%m.%Y', '%m/%d/%Y')):
    """Tries parsing a date string using multiple common formats."""
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val
    val_str = str(val).strip()
    for fmt in formats:
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Could not parse date: {val}")

@transaction.atomic
def ingest_sap_file(job_id, file_obj, filename):
    job = IngestionJob.objects.select_related('data_source', 'data_source__tenant').get(id=job_id)
    tenant = job.data_source.tenant
    config = job.data_source.config or {}
    plant_mapping = config.get('plant_mapping', {})

    job.status = 'PROCESSING'
    job.save()

    rows_to_create = []
    row_count = 0

    try:
        data_generator = read_file_data(file_obj, filename)
        for index, row in enumerate(data_generator, start=1):
            mvt = str(row.get('Bewegungsart', '')).strip()
            if mvt not in ('261', '201', '101'):
                continue
            
            row_count += 1
            flag_reasons = []
            
            qty_raw = row.get('Menge')
            parsed_qty = Decimal('0')
            try:
                parsed_qty = parse_german_decimal(qty_raw)
                if parsed_qty < 0:
                    flag_reasons.append("Negative quantity detected")
            except (InvalidOperation, ValueError):
                parsed_qty = Decimal('0')
                flag_reasons.append(f"Invalid quantity format: {qty_raw}")

            raw_unit = str(row.get('Basismengeneinheit', '')).strip()
            unit_map = {
                'L': 'liter', 'LITER': 'liter', 'LITERS': 'liter',
                'GAL': 'gallons', 'GALLON': 'gallons', 'GALLONS': 'gallons',
                'KG': 'kg', 'KILOGRAM': 'kg', 'KILOGRAMS': 'kg',
                'STK': 'pieces', 'PIECE': 'pieces', 'PIECES': 'pieces',
                'MWh': 'MWh', 'KWH': 'kWh', 'MI': 'miles', 'MILES': 'miles'
            }
            mapped_unit = unit_map.get(raw_unit.upper(), raw_unit)
            valid_units = ['liter', 'gallons', 'kg', 'pieces', 'MWh', 'kWh', 'miles', 'km', 'room-night']
            if mapped_unit not in valid_units:
                flag_reasons.append(f"Unrecognized unit: {raw_unit}")

            werk = str(row.get('Werk', '')).strip()
            if not werk:
                flag_reasons.append("Missing plant code (Werk)")
                location = "Unknown Plant"
            else:
                location = plant_mapping.get(werk)
                if not location:
                    location = f"Plant {werk}"
                    flag_reasons.append(f"Plant code {werk} not mapped to a location")

            post_date_raw = row.get('Buchungsdatum')
            activity_date = None
            try:
                activity_date = parse_standard_date(post_date_raw)
            except ValueError:
                flag_reasons.append(f"Could not parse posting date: {post_date_raw}")
                activity_date = timezone.now().date()

            material = str(row.get('Material', '')).strip()
            desc = str(row.get('Materialkurztext', '')).strip()
            
            is_fuel = any(x in desc.upper() or x in material.upper() for x in ['DIESEL', 'BENZIN', 'PETROL', 'FUEL', 'DI'])
            
            if is_fuel:
                source_type = 'SAP_FUEL'
                scope = 'SCOPE_1'
            else:
                source_type = 'SAP_PROCUREMENT'
                scope = 'SCOPE_3'

            status = 'FLAGGED' if flag_reasons else 'PENDING_REVIEW'
            activity_row = RawActivityRow(
                tenant=tenant,
                ingestion_job=job,
                source_type=source_type,
                scope=scope,
                raw_data=row,
                parsed_quantity=parsed_qty,
                parsed_unit=mapped_unit,
                activity_date=activity_date,
                location=location,
                description=desc or f"SAP Material {material}",
                status=status,
                flag_reasons=flag_reasons
            )
            rows_to_create.append(activity_row)

        RawActivityRow.objects.bulk_create(rows_to_create)
        
        audit_logs = []
        for r in RawActivityRow.objects.filter(ingestion_job=job):
            audit_logs.append(AuditLog(
                tenant=tenant,
                row=r,
                action='CREATED',
                performed_by=job.triggered_by,
                after_state={
                    'parsed_quantity': str(r.parsed_quantity),
                    'parsed_unit': r.parsed_unit,
                    'status': r.status,
                    'flag_reasons': r.flag_reasons
                }
            ))
        AuditLog.objects.bulk_create(audit_logs)

        job.status = 'COMPLETED'
        job.row_count = row_count
        job.error_count = len([r for r in rows_to_create if r.status == 'FLAGGED'])
        job.save()

        ds = job.data_source
        ds.last_ingested_at = timezone.now()
        ds.save()

    except Exception as e:
        job.status = 'FAILED'
        job.error_log = [{'error': str(e), 'step': 'Parsing file content'}]
        job.save()
        raise e

@transaction.atomic
def ingest_utility_file(job_id, file_obj, filename):
    job = IngestionJob.objects.select_related('data_source', 'data_source__tenant').get(id=job_id)
    tenant = job.data_source.tenant

    job.status = 'PROCESSING'
    job.save()

    rows_to_create = []
    row_count = 0

    try:
        data_generator = read_file_data(file_obj, filename)
        for index, row in enumerate(data_generator, start=1):
            row_count += 1
            flag_reasons = []

            p_start_raw = row.get('billing_period_start')
            p_end_raw = row.get('billing_period_end')
            
            p_start = None
            p_end = None
            activity_date = None

            try:
                p_start = parse_standard_date(p_start_raw)
                p_end = parse_standard_date(p_end_raw)
                if p_start and p_end:
                    days = (p_end - p_start).days
                    if days > 35:
                        flag_reasons.append(f"Billing period length is {days} days (exceeds 35 days)")
                    activity_date = p_start + timedelta(days=days // 2)
            except ValueError:
                flag_reasons.append(f"Invalid billing period dates: {p_start_raw} to {p_end_raw}")
                activity_date = timezone.now().date()

            cons_raw = row.get('consumption_kwh')
            parsed_qty = Decimal('0')
            try:
                parsed_qty = Decimal(str(cons_raw or '0').strip())
                if parsed_qty == 0:
                    flag_reasons.append("Zero energy consumption reported")
                elif parsed_qty < 0:
                    flag_reasons.append("Negative energy consumption detected")
            except (InvalidOperation, ValueError):
                flag_reasons.append(f"Invalid consumption value: {cons_raw}")

            raw_unit = str(row.get('consumption_unit', 'kWh')).strip()
            normalized_kwh = None
            if raw_unit.upper() == 'MWH':
                conv = UnitConversion.objects.filter(from_unit='MWh', to_unit='kWh').first()
                factor = conv.factor if conv else Decimal('1000')
                normalized_kwh = parsed_qty * factor
                mapped_unit = 'MWh'
            elif raw_unit.upper() == 'KWH':
                normalized_kwh = parsed_qty
                mapped_unit = 'kWh'
            else:
                flag_reasons.append(f"Unrecognized unit: {raw_unit}")
                mapped_unit = raw_unit

            account = row.get('account_number', 'N/A')
            meter = row.get('meter_id', 'N/A')
            site = row.get('site_name', 'N/A')
            supplier = row.get('supplier_name', 'N/A')

            status = 'FLAGGED' if flag_reasons else 'PENDING_REVIEW'
            activity_row = RawActivityRow(
                tenant=tenant,
                ingestion_job=job,
                source_type='UTILITY_ELECTRICITY',
                scope='SCOPE_2',
                raw_data=row,
                parsed_quantity=parsed_qty,
                parsed_unit=mapped_unit,
                normalized_quantity_kwh=normalized_kwh,
                activity_date=activity_date,
                period_start=p_start,
                period_end=p_end,
                location=f"Meter {meter} - Site {site}",
                description=f"Utility electricity invoice from {supplier} for account {account}",
                status=status,
                flag_reasons=flag_reasons
            )
            rows_to_create.append(activity_row)

        RawActivityRow.objects.bulk_create(rows_to_create)

        audit_logs = []
        for r in RawActivityRow.objects.filter(ingestion_job=job):
            audit_logs.append(AuditLog(
                tenant=tenant,
                row=r,
                action='CREATED',
                performed_by=job.triggered_by,
                after_state={
                    'parsed_quantity': str(r.parsed_quantity),
                    'parsed_unit': r.parsed_unit,
                    'status': r.status,
                    'flag_reasons': r.flag_reasons
                }
            ))
        AuditLog.objects.bulk_create(audit_logs)

        job.status = 'COMPLETED'
        job.row_count = row_count
        job.error_count = len([r for r in rows_to_create if r.status == 'FLAGGED'])
        job.save()

        ds = job.data_source
        ds.last_ingested_at = timezone.now()
        ds.save()

    except Exception as e:
        job.status = 'FAILED'
        job.error_log = [{'error': str(e), 'step': 'Parsing file content'}]
        job.save()
        raise e

@transaction.atomic
def ingest_travel_file(job_id, file_obj, filename):
    job = IngestionJob.objects.select_related('data_source', 'data_source__tenant').get(id=job_id)
    tenant = job.data_source.tenant

    job.status = 'PROCESSING'
    job.save()

    rows_to_create = []
    row_count = 0

    try:
        data_generator = read_file_data(file_obj, filename)
        for index, row in enumerate(data_generator, start=1):
            row_count += 1
            flag_reasons = []

            expense_type = str(row.get('expense_type', '')).strip()
            if not expense_type:
                flag_reasons.append("Missing expense type")
            amount_usd = Decimal('0')
            try:
                amount_raw = str(row.get('amount_usd', '0')).replace('$', '').replace(',', '').strip()
                amount_usd = Decimal(amount_raw)
                if amount_usd > 10000:
                    flag_reasons.append(f"Expense amount is high: {amount_usd} USD (exceeds 10,000 USD)")
            except (InvalidOperation, ValueError):
                flag_reasons.append(f"Invalid expense amount: {row.get('amount_usd')}")

            tx_date_raw = row.get('transaction_date')
            activity_date = None
            try:
                activity_date = parse_standard_date(tx_date_raw)
            except ValueError:
                flag_reasons.append(f"Could not parse transaction date: {tx_date_raw}")
                activity_date = timezone.now().date()

            origin = str(row.get('origin', '')).strip().upper()
            dest = str(row.get('destination', '')).strip().upper()
            employee = row.get('employee_name', 'Unknown')
            vendor = row.get('vendor_name', 'Unknown')

            parsed_qty = Decimal('0')
            mapped_unit = ''
            location = f"{origin} to {dest}" if origin or dest else "Various Locations"
            description = f"Corporate Travel - {expense_type} - {employee} via {vendor}"
            source_type = 'TRAVEL_GROUND'

            if expense_type.lower() == 'airfare':
                source_type = 'TRAVEL_FLIGHT'
                mapped_unit = 'km'
                
                if origin and dest and origin == dest:
                    flag_reasons.append(f"Flight origin and destination are identical: {origin}")
                
                dist_raw = row.get('distance_km')
                distance = Decimal('0')
                if dist_raw and str(dist_raw).strip() and str(dist_raw).strip() != '0':
                    try:
                        distance = Decimal(str(dist_raw).strip())
                    except (InvalidOperation, ValueError):
                        flag_reasons.append(f"Invalid custom distance: {dist_raw}")
                
                if distance == 0 and origin and dest and origin != dest:
                    coord1 = AIRPORT_COORDINATES.get(origin)
                    coord2 = AIRPORT_COORDINATES.get(dest)
                    if coord1 and coord2:
                        km = haversine_distance(coord1[0], coord1[1], coord2[0], coord2[1])
                        distance = Decimal(f"{km:.2f}")
                    else:
                        if not coord1:
                            flag_reasons.append(f"Airport code {origin} unrecognized in system database")
                        if not coord2:
                            flag_reasons.append(f"Airport code {dest} unrecognized in system database")
                        distance = Decimal('0')
                
                parsed_qty = distance

            elif expense_type.lower() == 'hotel':
                source_type = 'TRAVEL_HOTEL'
                mapped_unit = 'room-night'
                nights_raw = row.get('nights')
                try:
                    parsed_qty = Decimal(str(nights_raw or '0').strip())
                except (InvalidOperation, ValueError):
                    flag_reasons.append(f"Invalid number of nights: {nights_raw}")
                location = dest or origin or "Hotel Location"

            else:
                source_type = 'TRAVEL_GROUND'
                mapped_unit = 'km'
                dist_raw = row.get('distance_km')
                try:
                    if dist_raw and str(dist_raw).strip():
                        parsed_qty = Decimal(str(dist_raw).strip())
                    else:
                        parsed_qty = Decimal('0')
                        flag_reasons.append("Ground travel missing distance_km")
                except (InvalidOperation, ValueError):
                    flag_reasons.append(f"Invalid distance_km: {dist_raw}")

            status = 'FLAGGED' if flag_reasons else 'PENDING_REVIEW'
            activity_row = RawActivityRow(
                tenant=tenant,
                ingestion_job=job,
                source_type=source_type,
                scope='SCOPE_3',
                raw_data=row,
                parsed_quantity=parsed_qty,
                parsed_unit=mapped_unit,
                activity_date=activity_date,
                location=location,
                description=description,
                status=status,
                flag_reasons=flag_reasons
            )
            rows_to_create.append(activity_row)

        RawActivityRow.objects.bulk_create(rows_to_create)

        audit_logs = []
        for r in RawActivityRow.objects.filter(ingestion_job=job):
            audit_logs.append(AuditLog(
                tenant=tenant,
                row=r,
                action='CREATED',
                performed_by=job.triggered_by,
                after_state={
                    'parsed_quantity': str(r.parsed_quantity),
                    'parsed_unit': r.parsed_unit,
                    'status': r.status,
                    'flag_reasons': r.flag_reasons
                }
            ))
        AuditLog.objects.bulk_create(audit_logs)

        job.status = 'COMPLETED'
        job.row_count = row_count
        job.error_count = len([r for r in rows_to_create if r.status == 'FLAGGED'])
        job.save()

        ds = job.data_source
        ds.last_ingested_at = timezone.now()
        ds.save()

    except Exception as e:
        job.status = 'FAILED'
        job.error_log = [{'error': str(e), 'step': 'Parsing file content'}]
        job.save()
        raise e

def calculate_emissions(row):
    """
    Look up UnitConversion and EmissionFactor to perform calculations.
    Returns: (normalized_qty_kwh, normalized_qty_co2e, factor_used, factor_source, flag_reasons)
    """
    qty = row.parsed_quantity
    unit = row.parsed_unit
    source_type = row.source_type
    location = row.location or ""
    desc = row.description or ""
    
    normalized_kwh = None
    normalized_co2e = None
    factor_used = None
    factor_source = ""
    flags = list(row.flag_reasons or [])

    try:
        if source_type in ('SAP_FUEL', 'SAP_PROCUREMENT'):
            is_diesel = any(x in desc.upper() or x in row.description.upper() for x in ['DIESEL', 'DI'])
            if is_diesel:
                qty_liters = qty
                if unit == 'gallons':
                    conv = UnitConversion.objects.filter(from_unit='gallons', to_unit='liters').first()
                    if conv:
                        qty_liters = qty * conv.factor
                    else:
                        qty_liters = qty * Decimal('3.78541')
                elif unit != 'liter':
                    flags.append(f"Cannot auto-convert fuel from {unit} to liter")
                    return None, None, None, "", flags
                
                ef = EmissionFactor.objects.filter(activity_type='Diesel', region='UK').first()
                if ef:
                    factor_used = ef.factor_kg_co2e_per_unit
                    factor_source = f"{ef.source} {ef.valid_year}"
                    normalized_co2e = qty_liters * factor_used
                else:
                    factor_used = Decimal('2.68')
                    factor_source = "DEFRA 2023 (Fallback)"
                    normalized_co2e = qty_liters * factor_used
            else:
                ef = EmissionFactor.objects.filter(activity_type__iexact=desc).first()
                if not ef:
                    ef = EmissionFactor.objects.filter(activity_type__iexact=row.raw_data.get('Materialkurztext', '')).first()
                
                if ef:
                    factor_used = ef.factor_kg_co2e_per_unit
                    factor_source = f"{ef.source} {ef.valid_year}"
                    normalized_co2e = qty * factor_used
                else:
                    flags.append(f"No emission factor found for procurement material: {desc}")
                    normalized_co2e = Decimal('0')
                    factor_used = Decimal('0')
                    factor_source = "None"
        
        elif source_type == 'UTILITY_ELECTRICITY':
            qty_kwh = qty
            if unit == 'MWh':
                conv = UnitConversion.objects.filter(from_unit='MWh', to_unit='kWh').first()
                factor = conv.factor if conv else Decimal('1000')
                qty_kwh = qty * factor
            elif unit != 'kWh':
                conv = UnitConversion.objects.filter(from_unit=unit, to_unit='kWh').first()
                if conv:
                    qty_kwh = qty * conv.factor
                else:
                    flags.append(f"Cannot auto-convert electricity from {unit} to kWh")
                    return None, None, None, "", flags
            
            normalized_kwh = qty_kwh
            
            region = 'US'
            if 'UK' in location.upper() or 'UK' in desc.upper():
                region = 'UK'
                ef = EmissionFactor.objects.filter(activity_type='Grid electricity UK average', region='UK').first()
            else:
                ef = EmissionFactor.objects.filter(activity_type='Grid electricity US average', region='US').first()
                
            if ef:
                factor_used = ef.factor_kg_co2e_per_unit
                factor_source = f"{ef.source} {ef.valid_year}"
                normalized_co2e = qty_kwh * factor_used
            else:
                factor_used = Decimal('0.386') if region == 'US' else Decimal('0.207')
                factor_source = "EPA eGRID 2022 (Fallback)" if region == 'US' else "DEFRA 2023 (Fallback)"
                normalized_co2e = qty_kwh * factor_used

        elif source_type == 'TRAVEL_FLIGHT':
            qty_km = qty
            if unit == 'miles':
                conv = UnitConversion.objects.filter(from_unit='miles', to_unit='km').first()
                if conv:
                    qty_km = qty * conv.factor
                else:
                    qty_km = qty * Decimal('1.60934')
            
            booking_class = str(row.raw_data.get('booking_class', 'Y')).strip().upper()
            multipliers = {'Y': 1.0, 'W': 1.6, 'C': 2.0, 'F': 2.4}
            mult = multipliers.get(booking_class, 1.0)
            
            passenger_km = qty_km * Decimal(str(mult))
            
            if passenger_km < 3700:
                ef = EmissionFactor.objects.filter(activity_type='Short-haul flight').first()
                default_factor = Decimal('0.255')
            else:
                ef = EmissionFactor.objects.filter(activity_type='Long-haul flight').first()
                default_factor = Decimal('0.195')
                
            if ef:
                factor_used = ef.factor_kg_co2e_per_unit
                factor_source = f"{ef.source} {ef.valid_year}"
            else:
                factor_used = default_factor
                factor_source = "DEFRA 2023 (Fallback)"
                
            normalized_co2e = passenger_km * factor_used

        elif source_type == 'TRAVEL_HOTEL':
            ef = EmissionFactor.objects.filter(activity_type='Hotel stay').first()
            if ef:
                factor_used = ef.factor_kg_co2e_per_unit
                factor_source = f"{ef.source} {ef.valid_year}"
            else:
                factor_used = Decimal('31.9')
                factor_source = "DEFRA 2023 (Fallback)"
            normalized_co2e = qty * factor_used

        elif source_type == 'TRAVEL_GROUND':
            expense_type = str(row.raw_data.get('expense_type', '')).strip().lower()
            if 'car' in expense_type or 'rental' in expense_type:
                act_type = 'Car Rental'
                default_factor = Decimal('0.192')
            elif 'rail' in expense_type or 'train' in expense_type:
                act_type = 'Rail'
                default_factor = Decimal('0.035')
            else:
                act_type = 'Taxi'
                default_factor = Decimal('0.149')

            ef = EmissionFactor.objects.filter(activity_type=act_type).first()
            if ef:
                factor_used = ef.factor_kg_co2e_per_unit
                factor_source = f"{ef.source} {ef.valid_year}"
            else:
                factor_used = default_factor
                factor_source = "DEFRA 2023 (Fallback)"
            
            normalized_co2e = qty * factor_used
            
    except Exception as e:
        flags.append(f"Calculation error: {str(e)}")
        return None, None, None, "", flags

    return normalized_kwh, normalized_co2e, factor_used, factor_source, flags
